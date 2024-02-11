import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from time import sleep
from datetime import datetime, timedelta
from pathlib import Path
import sys
import pyautogui
from urllib.parse import quote
import webbrowser


def planinha_atualizada():
    os.system('cls')

    #Verifica se a pasta "planilha" existe caso não exista criará uma
    pasta = Path('Planilha')
    if not pasta.exists():
        pasta.mkdir()
        print('Pasta não encontrada')
        sleep(0.5)

        print('Criando nova pasta...')
        sleep(0.5)

        print('Pasta criada com sucesso!')
        sleep(0.5)
        pass


    diretorio = "Planilha/"

    # Lista todos os arquivos no diretório "Planilha"
    arquivos = os.listdir(diretorio)

    arquivo_xlsx = None
    
    for arquivo in arquivos:
        if arquivo.endswith(".xlsx"):
            arquivo_xlsx = os.path.join(diretorio, arquivo)
            break

    # Caso encontre algum arquivo ".xlsx" nas pasta "planilha" ele faz uma cópia das infomações desejadas da planilha encontrada 
    if arquivo_xlsx:
        print("Planilha encontrada!")
        sleep(0.5)
        
        print('Criando uma nova planilha atualizada...')
        sleep(0.5)

        workbook = openpyxl.load_workbook(f'{arquivo_xlsx}')
   
        pagina_clientes = workbook['Planilha1']

        wb = Workbook()
        ws = wb.active

        ws.append(['Nome', 'Número', 'Vencimento'])

        # Lista tudo que estiver na planilha antes de criar uma cópia
        for id, linha in enumerate(pagina_clientes.iter_rows(min_row=3)):
            
            
            print(f'{id}: {linha[1].value}| Número: {linha[15].value} | Vencimeto: {linha[25].value}')

       
        # escolhe as infomações que deseja passar para a planilha cópia
            nome = linha[1].value
            numero = linha[15].value
            vencimento = linha[25].value


            ws.append([f'{nome}', f'{numero}', f'{vencimento}'])

        sleep(2)
        
        os.system('cls')
        print('Planilha atualizada criada com sucesso!')
        wb.save('Planilha Atualizada.xlsx')
        sleep(1)


        # Carregando a planilha original e a planilha cópia
        wb_original = load_workbook(f'{arquivo_xlsx}')
        
        wb_copia = load_workbook('Planilha Atualizada.xlsx')

        # Selecionando as folhas ativas
        ws_original = wb_original.active
        ws_copia = wb_copia.active

        #Modifica o tamanho das colunas
        ws_copia.column_dimensions['A'].width = 40
        ws_copia.column_dimensions['B'].width = 20
        ws_copia.column_dimensions['C'].width = 15

        wb_copia.save('Planilha Atualizada.xlsx')


    else:
        # Caso não tenha uma planilha na pasta "planilha"
        os.system('cls')
        print('Nenhum Arquivo Foi Encontrado na Pasta')
        sleep(2)

        print('\nAdicione uma planilha nas pasta "planilha" para fazer uma cópia dos dados desejados e tente novamente.')

        os.system('cls')
        print('Fechando (3)')
        sleep(1)

        os.system('cls')
        print('Fechando (2)')
        sleep(1)

        os.system('cls')
        print('Fechando (1)')
        sleep(1)
        sys.exit()


def  menu():
    
    #Cria uma pasta "Não enviados" para armazenar planilhas com infomações das pessoas que não tiveram êxito ao enviar a mensagem
    pasta = Path('Não Enviados')
    if not pasta.exists():
        pasta.mkdir()

    planilha = Path('Não Enviados/Planilha de Reenvio.xlsx')

    # Verifica se a "Planinha de Reenvio" não exista, caso não exista cria uma
    if not planilha.exists():
        wb = Workbook()
        ws = wb.active

        ws.append(['Nome', 'Número', 'Vencimento'])

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

        wb.save('Não Enviados/Planilha de Reenvio.xlsx')
    
    while True:
        # Menu com as opções disponíveis para o bot
        os.system('cls')
        print('''WhatsApp Bot de Mensagem Automática

(1) Ativar Menssagens Automática
(2) Criar Lista de Clientes
              
                    ''')

        opcao = input('Digite 1 Para Ativar o Bot: ')

        # Caso a entrada seja "1" o bot é ativado
        if opcao == '1':
            sleep(2)
            mensagem_automatica()
            break
        # Caso a entrada seja "2" cria uma planilha cópia com as infomações desejadas
        if opcao == '2':
            planinha_atualizada()

        # caso seja qualquer outra coisa fora "1 e 2" não acontece nada
        else:
            continue

def mensagem_automatica():
    #função resposável pelo funcionamento do bot
    
    os.system('cls')

    # Abre a pagina do WhatsApp Web para verificar se está conectado ao mesmo. Após 20 segundo ele fecha e inicia o bot
    webbrowser.open('https://web.whatsapp.com/')
    sleep(20)
    pyautogui.hotkey('ctrl', 'w')
    
    workbook = openpyxl.load_workbook('Planilha Atualizada.xlsx')

    pagina_clientes = workbook['Sheet']

    # Extrai todos os dados da planilha cópia para o envio das mensagens
    for linha in pagina_clientes.iter_rows(min_row=2):

        nome = linha[0].value
        telefone = linha[1].value
        vencimento = linha[2].value

        # Caso o número seja vazio ele é alertado e registrado na planilha "Planilha de Reenvio"
        if telefone is None or telefone == '':
            
            print(f'Não foi possível enviar a mensagem para {nome} | (Sem Número)')

            workbook = load_workbook('Não Enviados/Planilha de Reenvio.xlsx')
            sheet = workbook.active
            
            linha = 1
            coluna = 1
            
            while sheet.cell(row=linha, column=coluna).value is not None:
                linha += 1

            # if sheet.cell(row=linha, column=coluna).value is None:
            dados = [f'{nome}', 'Sem Número', f'{vencimento}']
            
            for col, dado in enumerate(dados, start=1):
        
                sheet.cell(row=linha, column=col, value=dado)
            
            workbook.save('Não Enviados/Planilha de Reenvio.xlsx')
            continue
        

        data_antecipada = timedelta(days=int(vencimento)) - timedelta(days=1)        
        data_atual = datetime.now().day

        # Mensagem de exemplo que os clientes receberá contendo o nome e o vencimento. Sendo possível a troca a mensagem para a que mais agradar
        mensagem = f'''Mensagem Automática:
Olá {nome} seu boleto vence dia {vencimento}. Venha pagar presencialmente ou utilize nossos meios de pagamento:

Meios de pagamentos: ...

Caso o pagamento ja tenha sido efetuado, desconsidere esta mensagem.'''
        
        # Faz a verificação da data, caso seja um dia antes do vencimento ele enviará a mensagem, senão irá ignorar
        if data_antecipada.days == data_atual:

            link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'
            
            webbrowser.open(link_mensagem_whatsapp)
            sleep(15)
            
            # Localiza na tela a imagem para o envio do mensagem
            try:
                enviar = pyautogui.locateOnScreen('enviar.png')
                pyautogui.click(enviar[0], enviar[1])
                # pyautogui.click(1718, 1000)
                sleep(1)
                
                pyautogui.hotkey('ctrl', 'w')
            
            except Exception as error:
                
                # Caso não encontre a imagem na tela ele é alertado e registrado na "Planilha de Reenvio". Considerando que o número possívelmente seja inválido
                sleep(1)
                pyautogui.hotkey('ctrl', 'w')
                print(f'Não foi possível enviar a mensagem para {nome} | (Número Inválido)')
                
                workbook = load_workbook('Não Enviados/Planilha de Reenvio.xlsx')
                sheet = workbook.active
                
                linha = 1
                coluna = 1
                
                while sheet.cell(row=linha, column=coluna).value is not None:
                    linha += 1

                dados = [f'{nome}', 'Número Inválido', f'{vencimento}']
                
                for col, dado in enumerate(dados, start=1):
            
                    sheet.cell(row=linha, column=col, value=dado)
                
                workbook.save('Não Enviados/Planilha de Reenvio.xlsx')


    print('\nMensagens enviadas com sucesso!')
    input('Pressione ENTER para voltar')
    main()



def main():
    menu()

if __name__ == '__main__':
    main()